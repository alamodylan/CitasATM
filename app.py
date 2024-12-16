import re
from flask import Flask, render_template, request, redirect, url_for, send_file
from datetime import datetime, timedelta
import pytz
import os
from fpdf import FPDF
from openpyxl import Workbook
from flask import send_file
from flask import jsonify, render_template
from io import BytesIO
from openpyxl import Workbook
import psycopg2 
from psycopg2 import connect
from psycopg2.extras import RealDictCursor

app = Flask(__name__)


db_folder = "database"
db_path = os.path.join(db_folder, "citas.db")

if not os.path.exists(db_folder):
    os.makedirs(db_folder)

class PDF(FPDF):
    def header(self):
        self.image('static/LogoAlamo.png', 10, 8, 33) 
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'Alamo Terminales Marítimos - Detalle de Cita', border=False, ln=True, align='C')
        self.ln(20) 

    def footer(self):
        
        self.set_y(-30) 
        self.image('static/BASC.png', 10, self.get_y(), 33)  
        self.set_y(-25)  
        self.set_font('Arial', 'I', 8)
        self.multi_cell(0, 5, "Terminales - Transporte - Reparación - Certificado # CRSJO001581-1", align='C')

       
        self.set_y(-15)  
        self.cell(0, 10, f'Página {self.page_no()}', align='C')

def get_db_connection():
    DATABASE_URL = "postgresql://citasatm_user:SlwK1sFIPJal7m8KaDtlRlYu1NseKxnV@dpg-ctdis2jv2p9s73ai7op0-a.oregon-postgres.render.com/citasatm_db"
    try:
        # Intentar conectar a la base de datos
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
        return conn
    except Exception as e:
        # Imprimir el error y volver a lanzarlo
        print(f"Error al conectar con la base de datos: {e}")
        raise

def init_db():
    # Conexión a la base de datos PostgreSQL
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Crear la tabla 'citas' si no existe
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS citas (
                id SERIAL PRIMARY KEY,
                contenedor TEXT NOT NULL,
                chofer_nombre TEXT NOT NULL,
                chofer_cedula TEXT NOT NULL,
                cabezal_placa TEXT NOT NULL,
                fecha DATE NOT NULL,
                horario TEXT NOT NULL,
                naviera TEXT NOT NULL,
                estado_contenedor TEXT NOT NULL,  -- Nuevo campo
                tipo_operacion TEXT NOT NULL,     -- Nuevo campo
                estado TEXT DEFAULT 'Pendiente'
            )
        """)
        conn.commit()
        print("Tabla 'citas' creada o verificada exitosamente.")
    except Exception as e:
        print(f"Error al inicializar la base de datos: {e}")
    finally:
        cursor.close()
        conn.close()
@app.route("/")
def home():
    # Configurar la zona horaria
    zona_local = pytz.timezone('America/Costa_Rica')
    now = datetime.now(zona_local)  # Hora actual en zona horaria local

    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener las citas pendientes
        cursor.execute("SELECT * FROM citas WHERE estado = 'Pendiente'")
        pendientes = cursor.fetchall()

        # Verificar citas vencidas
        for cita in pendientes:
            cita_date = cita["fecha"]
            cita_end_time = cita["horario"].split("-")[1]

            # Convertir fecha y hora de la cita a zona horaria local
            cita_end_datetime = datetime.strptime(f"{cita_date} {cita_end_time}", "%Y-%m-%d %H:%M")
            cita_end_datetime = zona_local.localize(cita_end_datetime)

            # Comparar con la hora actual
            if cita_end_datetime + timedelta(hours=1) < now and cita["estado"] == "Pendiente":
                cursor.execute("UPDATE citas SET estado = 'Vencida' WHERE id = %s", (cita["id"],))
                conn.commit()

        # Actualizar pendientes
        cursor.execute("SELECT * FROM citas WHERE estado = 'Pendiente'")
        pendientes = cursor.fetchall()
    
    finally:
        cursor.close()
        conn.close()

    return render_template("index.html", citas=pendientes)

@app.route("/vencidas")
def vencidas():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Ejecutar la consulta para obtener citas vencidas
        cursor.execute("SELECT * FROM citas WHERE estado = 'Vencida'")
        vencidas = cursor.fetchall()
    finally:
        # Cerrar cursor y conexión
        cursor.close()
        conn.close()
    
    return render_template("vencidas.html", citas=vencidas)

@app.route("/completadas")
def completadas():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Ejecutar la consulta para obtener citas completadas
        cursor.execute("SELECT * FROM citas WHERE estado = 'Completada'")
        completadas = cursor.fetchall()
    finally:
        # Cerrar cursor y conexión
        cursor.close()
        conn.close()
    
    return render_template("completadas.html", citas=completadas)
@app.route("/revertir-cita/<int:id>", methods=["POST"])
def revertir_cita(id):
    codigo_autorizacion = request.json.get("codigo_autorizacion")

    # Validar el código de autorización
    if codigo_autorizacion != "atm7410":
        return "Código de autorización incorrecto.", 403

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cambiar el estado de la cita a "Pendiente"
        cursor.execute("UPDATE citas SET estado = %s WHERE id = %s", ('Pendiente', id))
        conn.commit()
    finally:
        # Cerrar cursor y conexión
        cursor.close()
        conn.close()

    return "", 204

@app.route("/guardar-anotacion/<int:id>", methods=["POST"])
def guardar_anotacion(id):
    anotacion = request.form.get("anotacion")
    print(f"Anotación recibida: {anotacion}")  # Para verificar qué se está recibiendo

    if not anotacion:
        print("Anotación vacía, retornando error")
        return "La anotación no puede estar vacía", 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Actualizar la anotación en la base de datos
        cursor.execute("UPDATE citas SET anotaciones = %s WHERE id = %s", (anotacion, id))
        conn.commit()
        print("Anotación guardada correctamente")
    except Exception as e:
        print(f"Error al guardar la anotación: {e}")
        return "Error al guardar la anotación", 500
    finally:
        # Cerrar cursor y conexión
        cursor.close()
        conn.close()

    return redirect(url_for("vencidas"))

@app.route("/crear-cita", methods=["GET", "POST"])
def crear_cita():
    if request.method == "POST":
        contenedor = request.form["contenedor"]
        chofer_nombre = request.form["chofer_nombre"]
        chofer_cedula = request.form["chofer_cedula"]
        cabezal_placa = request.form["cabezal_placa"]
        fecha = request.form["fecha"]
        horario = request.form["horario"]
        naviera = request.form["naviera"]
        estado_contenedor = request.form["estado_contenedor"]
        tipo_operacion = request.form["tipo_operacion"]

        # Validación del contenedor
        if not re.match(r"^[A-Z]{4}[0-9]{7}$", contenedor):
            return render_template("crear_cita.html", error="El contenedor debe contener 4 letras seguidas de 7 números (ejemplo: ABCD1234567)")

        # Validación de la naviera
        if naviera not in ["COSCO", "ONE", "OOCL"]:
            return render_template("crear_cita.html", error="La naviera debe ser COSCO, OOCL o ONE")

        # Validación del estado del contenedor
        if estado_contenedor not in ["Cargado", "Vacio"]:
            return render_template("crear_cita.html", error="El estado del contenedor debe ser 'Cargado' o 'Vacio'")

        # Validación del tipo de operación
        if tipo_operacion not in ["Retira", "Entrega"]:
            return render_template("crear_cita.html", error="El tipo de operación debe ser 'Retira' o 'Entrega'")

        # Validación del límite de citas por intervalo
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM citas
                WHERE fecha = %s AND horario = %s
            """, (fecha, horario))
            total_citas = cursor.fetchone()["total"]

            if total_citas >= 5:
                return render_template("crear_cita.html", error=f"El intervalo {horario} ya tiene el máximo de 5 citas.")

            # Insertar la cita
            cursor.execute("""
                INSERT INTO citas (contenedor, chofer_nombre, chofer_cedula, cabezal_placa, fecha, horario, naviera, estado_contenedor, tipo_operacion)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (contenedor, chofer_nombre, chofer_cedula, cabezal_placa, fecha, horario, naviera, estado_contenedor, tipo_operacion))
            conn.commit()
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for("home"))

    # Generar horarios dinámicos
    current_date = datetime.now().strftime("%Y-%m-%d")
    time_slots = []
    for hour in range(8, 17):  # Desde las 8:00 hasta las 17:00
        time_slots.extend([
            f"{hour:02d}:00-{hour:02d}:15",
            f"{hour:02d}:15-{hour:02d}:30",
            f"{hour:02d}:30-{hour:02d}:45",
            f"{hour:02d}:45-{(hour + 1):02d}:00"
        ])

    return render_template("crear_cita.html", time_slots=time_slots, current_date=current_date)

@app.route("/editar-cita/<int:id>", methods=["GET", "POST"])
def editar_cita(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM citas WHERE id = %s", (id,))
        cita = cursor.fetchone()

        if not cita:
            return render_template("error.html", message="Cita no encontrada.")

        # Generar horarios dinámicos
        time_slots = []
        for hour in range(8, 17):  # Desde las 8:00 hasta las 17:00
            time_slots.extend([
                f"{hour:02d}:00-{hour:02d}:15",
                f"{hour:02d}:15-{hour:02d}:30",
                f"{hour:02d}:30-{hour:02d}:45",
                f"{hour:02d}:45-{(hour + 1):02d}:00"
            ])

        if request.method == "POST":
            contenedor = request.form.get("contenedor")
            chofer_nombre = request.form.get("chofer_nombre")
            chofer_cedula = request.form.get("chofer_cedula")
            cabezal_placa = request.form.get("cabezal_placa")
            fecha = request.form.get("fecha")
            horario = request.form.get("horario")
            naviera = request.form.get("naviera")
            estado_contenedor = request.form.get("estado_contenedor")
            tipo_operacion = request.form.get("tipo_operacion")

            # Validación del contenedor
            if not re.match(r"^[A-Z]{4}[0-9]{7}$", contenedor):
                return render_template(
                    "editar_cita.html", 
                    cita=cita, 
                    time_slots=time_slots, 
                    error="El contenedor debe contener 4 letras seguidas de 7 números (ejemplo: ABCD1234567)"
                )

            # Validación de la naviera
            if naviera not in ["COSCO", "ONE", "OOCL"]:
                return render_template(
                    "editar_cita.html", 
                    cita=cita, 
                    time_slots=time_slots, 
                    error="La naviera debe ser COSCO, OOCL o ONE"
                )

            # Validación del estado del contenedor
            if estado_contenedor not in ["Cargado", "Vacio"]:
                return render_template(
                    "editar_cita.html", 
                    cita=cita, 
                    time_slots=time_slots, 
                    error="El estado del contenedor debe ser 'Cargado' o 'Vacio'"
                )

            # Validación del tipo de operación
            if tipo_operacion not in ["Retira", "Entrega"]:
                return render_template(
                    "editar_cita.html", 
                    cita=cita, 
                    time_slots=time_slots, 
                    error="El tipo de operación debe ser 'Retira' o 'Entrega'"
                )

            # Actualizar cita
            cursor.execute("""
                UPDATE citas
                SET contenedor = %s, chofer_nombre = %s, chofer_cedula = %s, cabezal_placa = %s, fecha = %s, horario = %s, naviera = %s, estado_contenedor = %s, tipo_operacion = %s
                WHERE id = %s
            """, (contenedor, chofer_nombre, chofer_cedula, cabezal_placa, fecha, horario, naviera, estado_contenedor, tipo_operacion, id))
            conn.commit()
            return redirect(url_for("home"))
    finally:
        cursor.close()
        conn.close()

    return render_template("editar_cita.html", cita=cita, time_slots=time_slots)


@app.route("/eliminar-cita/<int:id>", methods=["POST"])
def eliminar_cita(id):
    data = request.get_json()
    codigo_autorizacion = data.get("codigo_autorizacion")

    # Verificar el código de autorización
    if codigo_autorizacion != "12345":
        return "Código de autorización incorrecto.", 403

    # Eliminar la cita de la base de datos
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM citas WHERE id = %s", (id,))
        cita = cursor.fetchone()

        if not cita:
            return "Cita no encontrada.", 404

        cursor.execute("DELETE FROM citas WHERE id = %s", (id,))
        conn.commit()
        return "Cita eliminada con éxito.", 200
    finally:
        cursor.close()
        conn.close()

@app.route("/completar-cita/<int:id>", methods=["POST"])
def completar_cita(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Verificar si la cita existe
        cursor.execute("SELECT * FROM citas WHERE id = %s", (id,))
        cita = cursor.fetchone()
        if not cita:
            return render_template("error.html", message="Cita no encontrada.")

        # Verificar si la cita está pendiente
        if cita["estado"] != "Pendiente":
            return render_template("error.html", message="Solo se pueden completar citas pendientes.")

        # Cambiar el estado de la cita a "Completada"
        cursor.execute("UPDATE citas SET estado = 'Completada' WHERE id = %s", (id,))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for("home"))

@app.route("/generar-pdf/<int:cita_id>")
def generar_pdf(cita_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Obtener los detalles de la cita
        cursor.execute("SELECT * FROM citas WHERE id = %s", (cita_id,))
        cita = cursor.fetchone()

        if not cita:
            return render_template("error.html", message="Cita no encontrada.")

        # Crear el PDF
        pdf = PDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        pdf.cell(200, 10, txt="Detalles de la Cita", ln=True, align="C")
        pdf.ln(10)

        pdf.cell(0, 10, f"Contenedor: {cita['contenedor']}", ln=True)
        pdf.cell(0, 10, f"Chofer: {cita['chofer_nombre']}", ln=True)
        pdf.cell(0, 10, f"Cédula: {cita['chofer_cedula']}", ln=True)
        pdf.cell(0, 10, f"Placa: {cita['cabezal_placa']}", ln=True)
        pdf.cell(0, 10, f"Naviera: {cita['naviera']}", ln=True)
        pdf.cell(0, 10, f"Estado del Contenedor: {cita['estado_contenedor']}", ln=True)
        pdf.cell(0, 10, f"Tipo de Operación: {cita['tipo_operacion']}", ln=True)
        pdf.cell(0, 10, f"Fecha: {cita['fecha']}", ln=True)
        pdf.cell(0, 10, f"Horario: {cita['horario']}", ln=True)

        # Guardar el PDF en el directorio estático
        pdf_path = os.path.join("static", f"cita_{cita_id}.pdf")
        pdf.output(pdf_path)
    finally:
        cursor.close()
        conn.close()

    # Enviar el archivo PDF como descarga
    return send_file(pdf_path, as_attachment=True)

@app.route('/dashboard', methods=['GET'])
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener valores de filtros desde la solicitud
    fecha = request.args.get('fecha', '')
    naviera = request.args.get('naviera', '')
    tipo_operacion = request.args.get('tipo_operacion', '')
    horario = request.args.get('horario', '')
    estado = request.args.get('estado', '')
    estado_contenedor = request.args.get('estado_contenedor', '')  # Nuevo filtro

    # Construir consulta dinámica basada en los filtros
    query = "SELECT * FROM citas WHERE 1=1"
    params = []

    if fecha:
        query += " AND fecha = %s"
        params.append(fecha)
    if naviera:
        query += " AND naviera = %s"
        params.append(naviera)
    if tipo_operacion:
        query += " AND tipo_operacion = %s"
        params.append(tipo_operacion)
    if horario:
        query += " AND horario = %s"
        params.append(horario)
    if estado:
        query += " AND estado = %s"
        params.append(estado)
    if estado_contenedor:  # Aplica el nuevo filtro
        query += " AND estado_contenedor = %s"
        params.append(estado_contenedor)

    # Ejecutar la consulta con filtros
    cursor.execute(query, params)
    citas_filtradas = cursor.fetchall()

    # Estadísticas generales
    cursor.execute("SELECT COUNT(*) AS total FROM citas")
    total_citas = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT estado_contenedor, COUNT(*) AS total
        FROM citas
        GROUP BY estado_contenedor
    """)
    citas_por_estado = cursor.fetchall()

    cursor.execute("""
        SELECT DISTINCT horario
        FROM citas
        ORDER BY horario ASC
    """)
    horarios = cursor.fetchall()

    cursor.close()
    conn.close()

    # Renderizar el template con los datos
    return render_template(
        'dashboard.html',
        total_citas=total_citas,
        citas_por_estado=[dict(row) for row in citas_por_estado],
        horarios=[row["horario"] for row in horarios],
        citas_filtradas=[dict(row) for row in citas_filtradas],
        filtros={
            "fecha": fecha,
            "naviera": naviera,
            "tipo_operacion": tipo_operacion,
            "horario": horario,
            "estado": estado,
            "estado_contenedor": estado_contenedor,  # Pasa el nuevo filtro al template
        }
    )
@app.route("/exportar-estadisticas")
def exportar_estadisticas():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Recibir parámetros de filtro desde el formulario
    filtro_fecha = request.args.get("fecha")
    filtro_naviera = request.args.get("naviera")
    filtro_tipo_operacion = request.args.get("tipo_operacion")
    filtro_horario = request.args.get("horario")
    filtro_estado_contenedor = request.args.get("estado_contenedor")  # Nuevo filtro

    # Construir consulta con filtros dinámicos
    query = """
        SELECT *
        FROM citas
        WHERE 1=1
    """
    params = []

    if filtro_fecha:
        query += " AND fecha = %s"
        params.append(filtro_fecha)
    if filtro_naviera:
        query += " AND naviera = %s"
        params.append(filtro_naviera)
    if filtro_tipo_operacion:
        query += " AND tipo_operacion = %s"
        params.append(filtro_tipo_operacion)
    if filtro_horario:
        query += " AND horario = %s"
        params.append(filtro_horario)
    if filtro_estado_contenedor:  # Aplica el nuevo filtro
        query += " AND estado_contenedor = %s"
        params.append(filtro_estado_contenedor)

    # Ejecutar consulta
    cursor.execute(query, params)
    citas_filtradas = cursor.fetchall()

    cursor.close()
    conn.close()

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas Filtradas"

    # Encabezados
    ws.append(["ID", "Contenedor", "Chofer", "Naviera", "Estado Contenedor", "Tipo Operación", "Horario", "Fecha"])
    for cita in citas_filtradas:
        ws.append([
            cita["id"], cita["contenedor"], cita["chofer_nombre"], cita["naviera"],
            cita["estado_contenedor"], cita["tipo_operacion"], cita["horario"], cita["fecha"]
        ])

    # Guardar en memoria
    file_path = "estadisticas_filtradas.xlsx"
    wb.save(file_path)

    # Enviar archivo
    return send_file(file_path, as_attachment=True, download_name="estadisticas_filtradas.xlsx")

@app.route("/exportar-citas", methods=["GET"])
def exportar_citas():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Recuperar los filtros de la solicitud GET
    fecha = request.args.get('fecha', '')
    naviera = request.args.get('naviera', '')
    tipo_operacion = request.args.get('tipo_operacion', '')
    horario = request.args.get('horario', '')
    estado_contenedor = request.args.get('estado_contenedor', '')  # Nuevo filtro

    # Construir la consulta con los filtros
    query = """
        SELECT id, contenedor, chofer_nombre, naviera, estado_contenedor, tipo_operacion, fecha, horario, estado
        FROM citas
        WHERE 1=1
    """
    params = []

    if fecha:
        query += " AND fecha = %s"
        params.append(fecha)
    if naviera:
        query += " AND naviera = %s"
        params.append(naviera)
    if tipo_operacion:
        query += " AND tipo_operacion = %s"
        params.append(tipo_operacion)
    if horario:
        query += " AND horario = %s"
        params.append(horario)
    if estado_contenedor:  # Aplica el nuevo filtro
        query += " AND estado_contenedor = %s"
        params.append(estado_contenedor)

    # Ejecutar la consulta
    cursor.execute(query, params)
    citas = cursor.fetchall()

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas Filtradas"

    # Encabezados
    ws.append(["ID", "Contenedor", "Chofer", "Naviera", "Estado Contenedor", "Tipo Operación", "Fecha", "Horario", "Estado"])

    # Datos
    for cita in citas:
        ws.append([
            cita["id"], cita["contenedor"], cita["chofer_nombre"], cita["naviera"], cita["estado_contenedor"],
            cita["tipo_operacion"], cita["fecha"], cita["horario"], cita["estado"]
        ])

    file_path = "citas_filtradas.xlsx"
    wb.save(file_path)

    cursor.close()
    conn.close()
    return send_file(file_path, as_attachment=True, download_name="citas_filtradas.xlsx")

@app.route("/exportar-todas-citas", methods=["GET"])
def exportar_todas_citas():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener todas las citas de la base de datos
    cursor.execute("""
        SELECT id, contenedor, chofer_nombre, chofer_cedula, cabezal_placa, naviera, estado_contenedor, tipo_operacion, fecha, horario, estado
        FROM citas
    """)
    citas = cursor.fetchall()

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado de Citas"

    # Encabezados
    ws.append([
        "ID", "Contenedor", "Nombre del Chofer", "Cédula del Chofer", "Placa del Cabezal", 
        "Naviera", "Estado del Contenedor", "Tipo de Operación", "Fecha", "Horario", "Estado"
    ])

    # Agregar datos al archivo
    for cita in citas:
        ws.append([
            cita["id"], cita["contenedor"], cita["chofer_nombre"], cita["chofer_cedula"], 
            cita["cabezal_placa"], cita["naviera"], cita["estado_contenedor"], 
            cita["tipo_operacion"], cita["fecha"], cita["horario"], cita["estado"]
        ])

    # Guardar archivo en memoria temporal
    file_path = "todas_citas.xlsx"
    wb.save(file_path)

    # Cerrar cursor y conexión
    cursor.close()
    conn.close()

    # Enviar el archivo Excel al usuario
    return send_file(file_path, as_attachment=True, download_name="listado_citas_completo.xlsx")
DATABASE_URL = "postgresql://citasatm_user:SlwK1sFIPJal7m8KaDtlRlYu1NseKxnV@dpg-ctdis2jv2p9s73ai7op0-a.oregon-postgres.render.com/citasatm_db"



if __name__ == "__main__":
    if not os.path.exists(db_path):
        init_db()

    # Usa el puerto asignado por Render, o 5000 como respaldo
    port = int(os.environ.get("PORT", 5000))

    # Ejecuta la app con host accesible públicamente
    app.run(host="0.0.0.0", port=port)




