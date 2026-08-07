# services/reporte_excel_service.py

import os

from io import BytesIO

from datetime import datetime

import pytz

from flask import current_app

from openpyxl import Workbook

from openpyxl.drawing.image import Image

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter

from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo
)

from config import Config


# =========================================================
# COLORES CORPORATIVOS
# =========================================================
COLOR_AZUL = "003366"
COLOR_AZUL_CLARO = "D9EAF7"
COLOR_GRIS_CLARO = "F4F6F9"
COLOR_GRIS_BORDE = "D9DEE3"
COLOR_BLANCO = "FFFFFF"
COLOR_NEGRO = "212529"
COLOR_VERDE = "198754"
COLOR_AMARILLO_SUAVE = "FFF3CD"


# =========================================================
# CREAR EXCEL REPORTE MERCHANT
# =========================================================
def create_merchant_excel(
    registros,
    fecha_desde,
    fecha_hasta,
    naviera,
    username
):

    # =====================================================
    # CREAR LIBRO
    # =====================================================
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Reporte Merchant"


    # =====================================================
    # CONFIGURACIÓN GENERAL
    # =====================================================
    worksheet.sheet_view.showGridLines = False

    worksheet.freeze_panes = "A10"

    worksheet.page_setup.orientation = "landscape"

    worksheet.page_setup.paperSize = (
        worksheet.PAPERSIZE_A4
    )

    worksheet.page_setup.fitToWidth = 1

    worksheet.page_setup.fitToHeight = 0

    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.oddFooter.center.text = (
        "Álamo Terminales Marítimos"
    )

    worksheet.oddFooter.right.text = (
        "Página &P de &N"
    )

    worksheet.oddFooter.left.text = (
        "Reporte Merchant"
    )

    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.50
    worksheet.page_margins.bottom = 0.50


    # =====================================================
    # ESTILOS
    # =====================================================
    thin_border = Border(

        left=Side(
            style="thin",
            color=COLOR_GRIS_BORDE
        ),

        right=Side(
            style="thin",
            color=COLOR_GRIS_BORDE
        ),

        top=Side(
            style="thin",
            color=COLOR_GRIS_BORDE
        ),

        bottom=Side(
            style="thin",
            color=COLOR_GRIS_BORDE
        )

    )

    title_font = Font(
        name="Arial",
        size=18,
        bold=True,
        color=COLOR_AZUL
    )

    subtitle_font = Font(
        name="Arial",
        size=10,
        color="6C757D"
    )

    label_font = Font(
        name="Arial",
        size=10,
        bold=True,
        color=COLOR_AZUL
    )

    normal_font = Font(
        name="Arial",
        size=10,
        color=COLOR_NEGRO
    )

    header_font = Font(
        name="Arial",
        size=10,
        bold=True,
        color=COLOR_BLANCO
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=COLOR_AZUL
    )

    alternate_fill = PatternFill(
        fill_type="solid",
        fgColor=COLOR_GRIS_CLARO
    )

    porton_fill = PatternFill(
        fill_type="solid",
        fgColor=COLOR_AMARILLO_SUAVE
    )


    # =====================================================
    # LOGO
    # =====================================================
    logo_path = os.path.join(
        current_app.root_path,
        "static",
        "img",
        "LogoAlamo.png"
    )

    if os.path.exists(
        logo_path
    ):

        try:

            logo = Image(
                logo_path
            )

            logo.width = 115
            logo.height = 70

            worksheet.add_image(
                logo,
                "A1"
            )

        except Exception as error:

            print(
                "[REPORTE EXCEL] "
                f"No se pudo cargar logo: {error}"
            )


    # =====================================================
    # TÍTULO
    # =====================================================
    worksheet.merge_cells(
        "C1:H2"
    )

    title_cell = worksheet[
        "C1"
    ]

    title_cell.value = (
        "REPORTE MERCHANT"
    )

    title_cell.font = (
        title_font
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


    worksheet.merge_cells(
        "C3:H3"
    )

    subtitle_cell = worksheet[
        "C3"
    ]

    subtitle_cell.value = (
        "ÁLAMO TERMINALES MARÍTIMOS"
    )

    subtitle_cell.font = Font(
        name="Arial",
        size=11,
        bold=True,
        color="6C757D"
    )

    subtitle_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


    # =====================================================
    # DATOS DEL REPORTE
    # =====================================================
    zona_local = pytz.timezone(
        Config.TIMEZONE
    )

    fecha_generacion = datetime.now(
        zona_local
    )

    naviera_texto = (
        "Todas"
        if (
            not naviera
            or naviera == "TODAS"
        )
        else naviera
    )


    datos_reporte = [

        (
            "Fecha desde:",
            fecha_desde
        ),

        (
            "Fecha hasta:",
            fecha_hasta
        ),

        (
            "Naviera:",
            naviera_texto
        ),

        (
            "Generado por:",
            username or "Sistema"
        )

    ]


    # =====================================================
    # INFORMACIÓN IZQUIERDA
    # =====================================================
    row_info = 5

    for label, value in datos_reporte:

        worksheet.cell(
            row=row_info,
            column=1
        ).value = label

        worksheet.cell(
            row=row_info,
            column=1
        ).font = label_font

        worksheet.cell(
            row=row_info,
            column=2
        ).value = value

        worksheet.cell(
            row=row_info,
            column=2
        ).font = normal_font

        row_info += 1


    # =====================================================
    # INFORMACIÓN DERECHA
    # =====================================================
    worksheet[
        "F5"
    ] = "Fecha generación:"

    worksheet[
        "F5"
    ].font = label_font


    worksheet[
        "G5"
    ] = fecha_generacion

    worksheet[
        "G5"
    ].number_format = (
        "dd/mm/yyyy hh:mm"
    )

    worksheet[
        "G5"
    ].font = normal_font


    worksheet[
        "F6"
    ] = "Total registros:"

    worksheet[
        "F6"
    ].font = label_font


    worksheet[
        "G6"
    ] = len(
        registros
    )

    worksheet[
        "G6"
    ].font = Font(
        name="Arial",
        size=11,
        bold=True,
        color=COLOR_AZUL
    )


    # =====================================================
    # ENCABEZADOS DE TABLA
    # =====================================================
    header_row = 9

    headers = [
        "Contenedor",
        "BK / BL",
        "Fecha",
        "Servicio Terminal",
        "Chofer",
        "Placa",
        "Naviera",
        "Predio"
    ]


    for column_index, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header
        )

        cell.font = (
            header_font
        )

        cell.fill = (
            header_fill
        )

        cell.border = (
            thin_border
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    worksheet.row_dimensions[
        header_row
    ].height = 28


    # =====================================================
    # DATOS
    # =====================================================
    data_start_row = (
        header_row + 1
    )

    current_row = data_start_row


    for index, registro in enumerate(
        registros
    ):

        # =================================================
        # DETECTAR CONTENEDOR ORIGINAL / PORTÓN
        # =================================================
        contenedor_original = (
            registro.get(
                "contenedor"
            )
            if hasattr(
                registro,
                "get"
            )
            else None
        )

        contenedor_reporte = (
            registro.get(
                "contenedor_reporte",
                "-"
            )
            if hasattr(
                registro,
                "get"
            )
            else registro[
                "contenedor_reporte"
            ]
        )


        values = [

            contenedor_reporte,

            registro.get(
                "bk_bl",
                "-"
            ),

            registro.get(
                "fecha"
            ),

            registro.get(
                "servicio_terminal",
                "-"
            ),

            registro.get(
                "chofer_nombre",
                "-"
            ),

            registro.get(
                "cabezal_placa",
                "-"
            ),

            registro.get(
                "naviera",
                "-"
            ),

            registro.get(
                "predio_nombre",
                "-"
            )

        ]


        for column_index, value in enumerate(
            values,
            start=1
        ):

            cell = worksheet.cell(
                row=current_row,
                column=column_index,
                value=value
            )

            cell.font = normal_font

            cell.border = thin_border

            cell.alignment = Alignment(
                vertical="center"
            )


            # =============================================
            # FILAS ALTERNADAS
            # =============================================
            if index % 2 == 1:

                cell.fill = (
                    alternate_fill
                )


        # =================================================
        # FORMATO FECHA
        # =================================================
        worksheet.cell(
            row=current_row,
            column=3
        ).number_format = (
            "dd/mm/yyyy"
        )


        # =================================================
        # CONTENEDOR REGISTRADO EN PORTÓN
        # =================================================
        if (
            not contenedor_original
            and contenedor_reporte
            and contenedor_reporte != "-"
        ):

            worksheet.cell(
                row=current_row,
                column=1
            ).fill = porton_fill


        current_row += 1


    # =====================================================
    # TABLA EXCEL
    # =====================================================
    if registros:

        table_end_row = (
            current_row - 1
        )

        table_reference = (
            f"A{header_row}:"
            f"H{table_end_row}"
        )

        table = Table(
            displayName="TablaMerchant",
            ref=table_reference
        )

        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False
        )

        table.tableStyleInfo = (
            table_style
        )

        worksheet.add_table(
            table
        )


    # =====================================================
    # AUTOFILTRO
    # =====================================================
    if registros:

        worksheet.auto_filter.ref = (
            f"A{header_row}:"
            f"H{current_row - 1}"
        )


    # =====================================================
    # ANCHOS DE COLUMNAS
    # =====================================================
    column_widths = {

        "A": 20,  # Contenedor
        "B": 22,  # BK / BL
        "C": 14,  # Fecha
        "D": 20,  # Servicio Terminal
        "E": 32,  # Chofer
        "F": 16,  # Placa
        "G": 16,  # Naviera
        "H": 20   # Predio

    }

    for column, width in (
        column_widths.items()
    ):

        worksheet.column_dimensions[
            column
        ].width = width


    # =====================================================
    # ALINEACIONES
    # =====================================================
    for row in worksheet.iter_rows(
        min_row=data_start_row,
        max_row=max(
            data_start_row,
            current_row - 1
        ),
        min_col=1,
        max_col=8
    ):

        # Contenedor
        row[0].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # BK / BL
        row[1].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Fecha
        row[2].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Servicio Terminal
        row[3].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Chofer
        row[4].alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        # Placa
        row[5].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Naviera
        row[6].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Predio
        row[7].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    # =====================================================
    # ALTURA DE FILAS
    # =====================================================
    for row_number in range(
        data_start_row,
        current_row
    ):

        worksheet.row_dimensions[
            row_number
        ].height = 22


    # =====================================================
    # ÁREA DE IMPRESIÓN
    # =====================================================
    if registros:

        worksheet.print_area = (
            f"A1:H{current_row - 1}"
        )

    else:

        worksheet.print_area = (
            "A1:H10"
        )


    # =====================================================
    # REPETIR ENCABEZADO EN CADA PÁGINA
    # =====================================================
    worksheet.print_title_rows = (
        f"{header_row}:{header_row}"
    )


    # =====================================================
    # GUARDAR EN MEMORIA
    # =====================================================
    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(
        0
    )

    return output